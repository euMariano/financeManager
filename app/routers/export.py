"""Exportação das finanças para planilha."""
import io
from datetime import date

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlmodel import Session, select

from app.database import get_session
from app.models import Card, User
from app.routers.balance import get_or_create_balance
from app.security import get_current_user
from app.routers.cards import _refresh_card_percentages
from app.services import get_totals_and_zone

router = APIRouter(prefix="/export", tags=["export"])

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # verde claro
RESUMO_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")  # cinza claro
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

TYPE_LABELS = {
    "casa": "Casa",
    "faculdade": "Faculdade",
    "saude": "Saúde",
    "lazer": "Lazer",
    "alimentacao": "Alimentação",
    "transporte": "Transporte",
    "outros": "Outros",
}


def _build_workbook(session: Session, user: User) -> Workbook:
    balance = get_or_create_balance(session, user)
    cards = list(
        session.exec(
            select(Card).where(Card.user_id == user.id).order_by(Card.urgency, Card.due_date)
        ).all()
    )
    _refresh_card_percentages(session, cards, balance.net_balance)
    total_expenses, total_percentage, zone = get_totals_and_zone(cards, balance.net_balance)

    wb = Workbook()
    ws = wb.active
    ws.title = "Finanças"

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22

    row = 1
    ws[f"A{row}"] = "Organizador Financeiro"
    ws[f"A{row}"].font = Font(bold=True, size=14)
    ws.merge_cells(f"A{row}:B{row}")
    row += 2

    # --- Despesa por despesa (detalhamento) ---
    for i, card in enumerate(cards, 1):
        tipo_label = TYPE_LABELS.get(card.expense_type.value, card.expense_type.value)
        valor_br = f"{card.value:,.2f}".replace(",", " ").replace(".", ",").replace(" ", ".")
        nome = (getattr(card, "title", None) or "").strip() or "(sem título)"
        titulo = f"Despesa {i} — {nome} — {tipo_label} — R$ {valor_br}"

        ws[f"A{row}"] = titulo
        ws[f"A{row}"].font = Font(bold=True, size=11)
        ws[f"A{row}"].fill = SECTION_FILL
        ws.merge_cells(f"A{row}:B{row}")
        for c in ("A", "B"):
            ws[f"{c}{row}"].border = THIN_BORDER
        row += 1

        pct = (card.percentage or 0) / 100 if isinstance(card.percentage, (int, float)) else 0
        due = card.due_date.isoformat() if isinstance(card.due_date, date) else str(card.due_date)
        status = card.status.value.capitalize()

        titulo_valor = (getattr(card, "title", None) or "").strip()
        linhas = [
            ("Título", titulo_valor),
            ("Urgência (prioridade)", card.urgency),
            ("Tipo de despesa", tipo_label),
            ("Valor (R$)", round(card.value, 2)),
            ("% do saldo líquido", pct),
            ("Data para pagar", due),
            ("Status", status),
        ]
        for label, valor in linhas:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = valor
            ws[f"A{row}"].border = THIN_BORDER
            ws[f"B{row}"].border = THIN_BORDER
            if label == "Valor (R$)":
                ws[f"B{row}"].number_format = '"R$ "#,##0.00'
            elif label == "% do saldo líquido":
                ws[f"B{row}"].number_format = "0.00%"
            row += 1

        row += 1  # espaço entre despesas

    # --- Resumo geral (no fim) ---
    ws[f"A{row}"] = "Resumo geral"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].fill = RESUMO_FILL
    ws.merge_cells(f"A{row}:B{row}")
    for c in ("A", "B"):
        ws[f"{c}{row}"].border = THIN_BORDER
    row += 1

    ws[f"A{row}"] = "Saldo líquido total"
    ws[f"B{row}"] = round(balance.net_balance, 2)
    ws[f"B{row}"].number_format = '"R$ "#,##0.00'
    ws[f"A{row}"].border = THIN_BORDER
    ws[f"B{row}"].border = THIN_BORDER
    row += 1

    ws[f"A{row}"] = "Total de todas as despesas"
    ws[f"B{row}"] = round(total_expenses, 2)
    ws[f"B{row}"].number_format = '"R$ "#,##0.00'
    ws[f"A{row}"].border = THIN_BORDER
    ws[f"B{row}"].border = THIN_BORDER
    row += 1

    ws[f"A{row}"] = "Percentual total (despesas sobre saldo)"
    ws[f"B{row}"] = round(total_percentage, 2) / 100
    ws[f"B{row}"].number_format = "0.00%"
    ws[f"A{row}"].border = THIN_BORDER
    ws[f"B{row}"].border = THIN_BORDER
    row += 1

    ws[f"A{row}"] = "Situação (faixa)"
    ws[f"B{row}"] = zone.value.upper()
    ws[f"A{row}"].border = THIN_BORDER
    ws[f"B{row}"].border = THIN_BORDER

    return wb


@router.get("/spreadsheet")
def export_spreadsheet(
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """Gera e retorna um arquivo Excel (.xlsx) com resumo e lista de despesas."""
    wb = _build_workbook(session, current_user)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=financas.xlsx"},
    )
